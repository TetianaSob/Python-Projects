# gui.py

from tkinter import *
from openpyxl import *

wb = load_workbook(r'data.xlsx')
sheet = wb.active

def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.cell(row=1,column=1).value = "Name"

def clear():
    name_field.delete(0, END)

def insert():

    if (name_field.get() == ""):
        print("empty input")
    
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        wb.save(r'data.xsx')
        clear()

if  __name__ == "__name__":
    root =Tk()
    root.title("registration form")
    excel()

    #create a Name label
    name = Label(root, text="Name")
    name.grid(row=1, column=0)

    # create a text entry box
    # for typing the information

    name_field = entry(root)

    name_field.grid(row=1, column=1, ipadx="100")

    #call excel function
    excel()

    submit =Button(root, text="Submit", command=insert)
    submit.grid(row=8, column=1)

    root.mainloop()
